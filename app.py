import os
import re
import sys
import json
import time
import uuid
import tempfile
import shutil
import webbrowser
import urllib.parse
from datetime import datetime, timedelta
import pythoncom
import openpyxl
import win32com.client
from flask import Flask, render_template, request, jsonify, send_file

# Configuracion de rutas (compatible con PyInstaller)
if getattr(sys, 'frozen', False):
    BUNDLE_DIR = sys._MEIPASS
    DIRECTORIO_ACTUAL = os.path.dirname(sys.executable)
else:
    BUNDLE_DIR = os.path.dirname(os.path.abspath(__file__))
    DIRECTORIO_ACTUAL = BUNDLE_DIR

ARCHIVO_DB = os.path.join(DIRECTORIO_ACTUAL, "Base_Clientes.xlsx")
HISTORIAL_FILE = os.path.join(DIRECTORIO_ACTUAL, "historial.json")
UPLOAD_FOLDER = os.path.join(DIRECTORIO_ACTUAL, "uploads")

app = Flask(__name__,
            template_folder=os.path.join(BUNDLE_DIR, 'templates'),
            static_folder=os.path.join(BUNDLE_DIR, 'static'))
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def cargar_clientes():
    if not os.path.exists(ARCHIVO_DB):
        return []
    try:
        wb = openpyxl.load_workbook(ARCHIVO_DB, data_only=True, read_only=True)
    except PermissionError:
        temp_path = os.path.join(UPLOAD_FOLDER, "_temp_clientes.xlsx")
        shutil.copy2(ARCHIVO_DB, temp_path)
        wb = openpyxl.load_workbook(temp_path, data_only=True, read_only=True)

    ws = wb.active
    clientes = []

    def safe_cell(row, idx):
        if idx >= len(row) or row[idx] is None:
            return ""
        val = row[idx]
        # Normalizar IDs numericos: 1.0 -> "1"
        if isinstance(val, (int, float)):
            if val == int(val):
                return str(int(val))
        return str(val)

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if not row or len(row) < 2 or not row[1]:
            continue
        clientes.append({
            "id": idx,
            "codigo": safe_cell(row, 0),
            "cliente": safe_cell(row, 1),
            "para": safe_cell(row, 2),
            "cc": safe_cell(row, 3),
            "asunto": safe_cell(row, 4).replace("\n", " "),
            "cuerpo": safe_cell(row, 5).replace("\n", "<br>")
        })
    wb.close()
    return clientes




def limpiar_parrafos_vacios(html):
    """Elimina <p class=MsoNormal> vacios que Outlook pone antes de la firma."""
    patron_p = re.compile(r'<p\b[^>]*>.*?</p>\s*', re.IGNORECASE | re.DOTALL)

    resultado = html
    for _ in range(10):
        match = patron_p.match(resultado)
        if not match:
            break
        texto = re.sub(r'<[^>]*>', '', match.group(0))
        texto = texto.replace('&nbsp;', '').replace('\r', '').replace('\n', '').strip()
        if texto:
            break
        resultado = resultado[match.end():]

    return resultado


# ── Historial ─────────────────────────────────────────────────────

def cargar_historial():
    if not os.path.exists(HISTORIAL_FILE):
        return []
    try:
        with open(HISTORIAL_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return []


def guardar_historial(entrada):
    historial = cargar_historial()
    historial.insert(0, entrada)  # Mas reciente primero
    with open(HISTORIAL_FILE, "w", encoding="utf-8") as f:
        json.dump(historial, f, ensure_ascii=False, indent=2)


# ── Rutas ─────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/clientes", methods=["GET"])
def api_clientes():
    return jsonify(cargar_clientes())


@app.route("/api/abrir_excel", methods=["GET"])
def abrir_excel():
    try:
        os.startfile(ARCHIVO_DB)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/generar_correo", methods=["POST"])
def generar_correo():
    try:
        data = request.form
        cliente_id = data.get("cliente_id")
        modo = data.get("modo", "clasico")  # "clasico" o "nuevo"

        clientes = cargar_clientes()
        cliente = next((c for c in clientes if str(c["id"]) == str(cliente_id)), None)

        if not cliente:
            return jsonify({"success": False, "error": "Cliente no encontrado"}), 404

        asunto = cliente.get("asunto", "").replace("[CLIENTE]", cliente["cliente"])
        cuerpo = cliente.get("cuerpo", "").replace("[CLIENTE]", cliente["cliente"])
        cuerpo = re.sub(r'(<br\s*/?>|\s)+$', '', cuerpo, flags=re.IGNORECASE)

        # ── Modo: Outlook Nuevo (mailto:) ─────────────────────
        if modo == "nuevo":
            para = cliente.get("para", "")
            cc = cliente.get("cc", "")
            # Limpiar HTML del cuerpo para mailto (texto plano)
            cuerpo_plano = re.sub(r'<br\s*/?>', '\n', cuerpo, flags=re.IGNORECASE)
            cuerpo_plano = re.sub(r'<[^>]*>', '', cuerpo_plano)

            params = {"subject": asunto, "body": cuerpo_plano}
            if cc:
                params["cc"] = cc

            mailto_url = f"mailto:{para}?{urllib.parse.urlencode(params, quote_via=urllib.parse.quote)}"

            # Guardar adjuntos en carpeta del Escritorio
            archivos = request.files.getlist("archivo")
            nombres_archivos = [f.filename for f in archivos if f and f.filename]

            if nombres_archivos:
                desktop = os.path.join(os.path.expanduser("~"), "Desktop")
                if not os.path.exists(desktop):
                    desktop = os.path.join(os.path.expanduser("~"), "Escritorio")
                carpeta_adjuntos = os.path.join(desktop, "Adjuntos_AquaShield")
                os.makedirs(carpeta_adjuntos, exist_ok=True)
                for archivo in archivos:
                    if archivo and archivo.filename:
                        archivo.save(os.path.join(carpeta_adjuntos, archivo.filename))
                # Abrir la carpeta para que el usuario arrastre al correo
                os.startfile(carpeta_adjuntos)

            # Abrir el correo
            webbrowser.open(mailto_url)

            msg = "Correo abierto en Outlook Nuevo."
            if nombres_archivos:
                msg += f" Los {len(nombres_archivos)} archivo(s) están en la carpeta 'Adjuntos_AquaShield' del Escritorio. Arrástralos al correo."

            # Registrar en historial
            guardar_historial({
                "timestamp": datetime.now().isoformat(),
                "cliente_id": cliente_id,
                "cliente": cliente["cliente"],
                "para": cliente.get("para", ""),
                "asunto": asunto,
                "archivo": ", ".join(nombres_archivos) if nombres_archivos else "",
                "modo": "nuevo"
            })

            return jsonify({"success": True, "message": msg, "modo": "nuevo"})

        # ── Modo: Outlook Clasico (COM) ───────────────────────
        archivos = request.files.getlist("archivo")
        rutas_archivos = []
        nombres_archivos = []
        # Usar carpeta TEMP del sistema (evita bloqueos de OneDrive)
        subfolder = os.path.join(tempfile.gettempdir(), "AquaShield_" + uuid.uuid4().hex)
        os.makedirs(subfolder, exist_ok=True)
        for archivo in archivos:
            if archivo and archivo.filename:
                ruta = os.path.join(subfolder, archivo.filename)
                archivo.save(ruta)
                rutas_archivos.append(ruta)
                nombres_archivos.append(archivo.filename)

        pythoncom.CoInitialize()
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)

            def formatear_correos(correos_str):
                if not correos_str:
                    return ""
                lista = [e.strip() for e in re.split(r'[,;]+', correos_str) if e.strip()]
                return "; ".join(lista)

            mail.Display()

            mail.To = formatear_correos(cliente.get("para", ""))
            mail.CC = formatear_correos(cliente.get("cc", ""))
            mail.Subject = asunto

            # Capturar HTML de Outlook (firma)
            firma_html = mail.HTMLBody
            cuerpo_html = "<p class=MsoNormal style='margin:0;padding:0'><span style='font-family:Calibri,sans-serif;font-size:11.0pt'>" + cuerpo + "</span></p><br>"

            # Inyectar dentro de WordSection1, limpiando parrafos vacios
            ws_match = re.search(r'<div\s+class=\s*["\']?WordSection1["\']?\s*>', firma_html, re.IGNORECASE)
            if ws_match:
                antes = firma_html[:ws_match.end()]
                despues = firma_html[ws_match.end():]
                despues = limpiar_parrafos_vacios(despues)
                mail.HTMLBody = antes + cuerpo_html + despues
            else:
                body_match = re.search(r'<body[^>]*>', firma_html, re.IGNORECASE)
                if body_match:
                    antes = firma_html[:body_match.end()]
                    despues = firma_html[body_match.end():]
                    despues = limpiar_parrafos_vacios(despues)
                    mail.HTMLBody = antes + cuerpo_html + despues
                else:
                    mail.HTMLBody = cuerpo_html + firma_html

            # Adjuntar todos los archivos
            for ruta in rutas_archivos:
                ruta_abs = os.path.abspath(ruta)
                print(f"[ADJUNTO] Adjuntando: {ruta_abs}")
                mail.Attachments.Add(ruta_abs)

            # Esperar a que Outlook procese los adjuntos antes de limpiar
            time.sleep(3)

        finally:
            pythoncom.CoUninitialize()

        # Limpiar archivos temporales
        for ruta in rutas_archivos:
            try:
                os.remove(ruta)
            except Exception as e:
                print(f"[WARN] No se pudo limpiar {ruta}: {e}")
        try:
            os.rmdir(subfolder)
        except Exception:
            pass

        # Registrar en historial
        guardar_historial({
            "timestamp": datetime.now().isoformat(),
            "cliente_id": cliente_id,
            "cliente": cliente["cliente"],
            "para": cliente.get("para", ""),
            "asunto": asunto,
            "archivo": ", ".join(nombres_archivos) if nombres_archivos else "",
            "modo": "clasico"
        })

        return jsonify({"success": True, "message": f"Correo abierto con {len(rutas_archivos)} adjunto(s)", "modo": "clasico"})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/historial", methods=["GET"])
def api_historial():
    return jsonify(cargar_historial())


@app.route("/api/stats", methods=["GET"])
def api_stats():
    historial = cargar_historial()
    ahora = datetime.now()
    hoy = ahora.date()
    inicio_semana = hoy - timedelta(days=hoy.weekday())
    inicio_mes = hoy.replace(day=1)

    correos_hoy = 0
    correos_semana = 0
    correos_mes = 0
    clientes_conteo = {}

    for entry in historial:
        try:
            fecha = datetime.fromisoformat(entry["timestamp"]).date()
        except (ValueError, KeyError):
            continue
        if fecha == hoy:
            correos_hoy += 1
        if fecha >= inicio_semana:
            correos_semana += 1
        if fecha >= inicio_mes:
            correos_mes += 1
        nombre = entry.get("cliente", "Desconocido")
        clientes_conteo[nombre] = clientes_conteo.get(nombre, 0) + 1

    # Top 3 clientes
    top_clientes = sorted(clientes_conteo.items(), key=lambda x: x[1], reverse=True)[:3]

    return jsonify({
        "hoy": correos_hoy,
        "semana": correos_semana,
        "mes": correos_mes,
        "total": len(historial),
        "top_clientes": [{"nombre": c[0], "cantidad": c[1]} for c in top_clientes]
    })


@app.route("/api/excel_timestamp", methods=["GET"])
def excel_timestamp():
    try:
        if os.path.exists(ARCHIVO_DB):
            mtime = os.path.getmtime(ARCHIVO_DB)
            return jsonify({"timestamp": mtime})
        return jsonify({"timestamp": 0})
    except Exception:
        return jsonify({"timestamp": 0})


@app.route("/api/exportar_historial", methods=["GET"])
def exportar_historial():
    historial = cargar_historial()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial de Correos"
    ws.append(["Fecha", "Cliente", "Destinatario", "Asunto", "Adjunto", "Modo"])

    for entry in historial:
        try:
            fecha = datetime.fromisoformat(entry["timestamp"]).strftime("%d/%m/%Y %H:%M")
        except (ValueError, KeyError):
            fecha = ""
        ws.append([
            fecha,
            entry.get("cliente", ""),
            entry.get("para", ""),
            entry.get("asunto", ""),
            entry.get("archivo", ""),
            entry.get("modo", "")
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    export_path = os.path.join(UPLOAD_FOLDER, "historial_correos.xlsx")
    wb.save(export_path)
    return send_file(export_path, as_attachment=True, download_name="Historial_Correos.xlsx")


if __name__ == "__main__":
    print("Iniciando Agente de Correos en http://localhost:5055")
    app.run(port=5055, debug=True)